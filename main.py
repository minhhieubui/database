import sys
import cv2
import os
import glob
import matplotlib.pyplot as plt
import numpy as np
import mysql.connector
import time
import datetime as dt

from PyQt5.QtGui import QFont, QImage, QPixmap
from PyQt5.QtWidgets import QApplication, QWidget, QMessageBox, QMainWindow, QCheckBox, QTableWidgetItem, QPushButton
from login import Ui_Form           # loin
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtMultimedia import QCamera
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSignal, QObject, QTimer
from PyQt5.QtWidgets import QFrame, QLineEdit, QButtonGroup
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from gui_windown import Ui_Gui_windown              # gui_windown
from a_few_location import Ui_Form_1                # a_few_location
from PyQt5 import QtCore
from datetime import datetime

from product import Ui_Form_2                       # product_info
from connect_database import ConnectDatabase
from openpyxl import Workbook


db = mysql.connector.connect(user='root', password='123456', host='127.0.0.1', database='data_login')
mycursor = db.cursor()

class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.uic1 = Ui_Form()  # Tạo đối tượng UI
        self.uic1.setupUi(self)  # Thiết lập giao diện

        self.setWindowFlag(Qt.FramelessWindowHint)  # Thiết lập cửa sổ không có viền
        self.setAttribute(Qt.WA_TranslucentBackground)  # Thiết lập nền của cửa sổ trong suốt

        # Kết nối nút nhấn Login, Register
        self.uic1.login_button.clicked.connect(self.login)
        self.uic1.toolButton_sign_up.clicked.connect(self.register)
        self.uic1.toolButton_account.clicked.connect(lambda: self.uic1.Background.setCurrentIndex(0))  # Nút Click Here Sign Up được ấn chuyển qua cửa sổ 2( Register)
        self.uic1.toolButton_clik_for_sign.clicked.connect(lambda: self.uic1.Background.setCurrentIndex(1))  # Nút Already Have a Account được ấn chuyển qua cửa sổ 1(Login)

        # Kết nối nút nhấn
        self.uic1.pushButton_close.clicked.connect(self.close)
        self.uic1.pushButton_close_1.clicked.connect(self.close)
        self.uic1.checkBox_show_pass.setChecked(False)  # Khởi tạo btn show_pass ban đầu = false
        self.uic1.checkBox_show_pass.stateChanged.connect(self.show_pass)  # Kiểm tra trạng thái thay đổi của btn show_pass

        lst = ['Admin', 'User']
        for x in lst:
            self.uic1.comboBox_admin.addItem(x)
        self.uic1.line_user.setFocus()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Down:
            self.uic1.line_password.setFocus()
        if event.key() == Qt.Key_Up:
            self.uic1.line_user.setFocus()
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:  # Khi nhấn nút enter
            self.login()

    def show_pass(self):
        if self.uic1.checkBox_show_pass.isChecked():
            self.uic1.line_password.setEchoMode(QLineEdit.Normal)  # Hiện password
        else:
            self.uic1.line_password.setEchoMode(QLineEdit.Password)  # Không hiện mật khẩu

    def login(self):
        try:
            self.uic1.msg = QtWidgets.QMessageBox()
            self.uic1.msg.setWindowTitle('Warning')

            username = self.uic1.line_user.text()
            password = self.uic1.line_password.text()

            print("User", username)
            print("Password", password)

            # Tạo table login trong data_login
            query = "SELECT * FROM login WHERE BINARY Username=%s AND Password=%s"
            mycursor.execute(query, (username, password))
            result = mycursor.fetchone()
            if result:
                self.close()
                global check_in
                if result[5] == "Admin":
                    check_in = "Admin"
                else:
                    check_in = "User"

                # datetime object containing current date and time
                now = datetime.now()
                # Tách ngày và giờ
                date = now.date()
                time = now.time()
                print(type(date))
                print("date and time =", date)
                query2 = "INSERT INTO history_access (Date, Time, Username, Role) VALUES (%s, %s, %s, %s)"
                mycursor.execute(query2, (date, time, result[2], result[5]))
                db.commit()

                self.main_win1 = MainWindow()
                self.main_win1.show()
            else:
                QMessageBox.information(self, "Message", "Incorrect Username or Password")
        except Exception as e:
            print("Error during login:", e)

    def register(self):
        try:
            fullname = self.uic1.line_fullname.text()
            username = self.uic1.line_user_1.text()
            email = self.uic1.line_email.text()
            password = self.uic1.line_password_1.text()
            role = self.uic1.comboBox_admin.currentText()

            query = "SELECT * FROM login WHERE Username=%s AND Password=%s AND Email=%s AND Fullname=%s AND Role=%s"
            mycursor.execute(query, (username, password, email, fullname, role))
            result = mycursor.fetchone()

            if result:
                self.uic1.msg = QtWidgets.QMessageBox()
                self.uic1.msg.setWindowTitle('Warning')
                self.uic1.msg.setText("Account Already Exists")
                self.uic1.msg.exec_()
            else:
                query1 = "INSERT INTO login (Fullname, Username, Email, Password, Role) VALUES (%s, %s, %s, %s, %s)"
                mycursor.execute(query1, (fullname, username, email, password, role))
                self.uic1.line_fullname.clear()
                self.uic1.line_user_1.clear()
                self.uic1.line_email.clear()
                self.uic1.line_password_1.clear()
                db.commit()
                self.uic1.Background.setCurrentIndex(0)
                self.uic1.msg = QtWidgets.QMessageBox()
                self.uic1.msg.setWindowTitle('')
                self.uic1.msg.setText("You have successfully registered")
                self.uic1.msg.exec_()
        except Exception as e:
            print("Error during registration:", e)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.uic = Ui_Gui_windown()  # Tạo đối tượng UI
        self.uic.setupUi(self)  # Thiết lập giao diện

        self.thread = {}  # Biến thread thông thường được tạo ra bằng rỗng
        self.init_button() # Khai báo nút nhấn
    def init_button(self):
        # Khai báo nút nhấn
        self.uic.pushButton_start_camera.clicked.connect(self.start_capture_video)
        self.uic.pushButton_stop_camera.clicked.connect(self.stop_capture_video)
        self.uic.pushButton_home_1.clicked.connect(self.start_home)

        self.uic.pushButton_total_location.clicked.connect(self.show_map)
        self.uic.pushButton_a_few_location.clicked.connect(self.choose_Ui_Form_1)

        self.uic.pushButton_auto_run.clicked.connect(self.run_drone)
        self.uic.pushButton_auto_stop.clicked.connect(self.stop_drone)
        self.uic.pushButton_reset_auto.clicked.connect(self.reset)

        self.uic.pushButton_start_mapping.clicked.connect(self.start_mapping)
        self.uic.pushButton_stop_mapping.clicked.connect(self.stop_mapping)

        # self.uic.pushButton_log_out.clicked.connect(self.log_out)
        self.uic.pushButton_search.clicked.connect(self.search)
        self.uic.pushButton_clear.clicked.connect(self.clear)

        self.uic.pushButton_calendar_1.clicked.connect(self.calendar1)
        self.uic.pushButton_calendar_2.clicked.connect(self.calendar2)

        self.uic.pushButton_load.clicked.connect(self.load_data)
        self.uic.pushButton_report.clicked.connect(self.report)



        global check_in
        if(check_in=="Admin"):
            self.uic.history.setEnabled(True)
            self.uic.main.setEnabled(True)
        else:
            self.uic.history.setEnabled(False)

        # Khai báo Calendar 1
        self.lich1 = QtWidgets.QCalendarWidget(self)  # Khởi tạo màn hình Calendar1 khi ấn nút
        self.lich1.setWindowTitle("Calendar")
        self.lich1.move(1100, 300)
        self.lich1.resize(400,300)
        self.lich1.clicked[QtCore.QDate].connect(self.get_data1)
        self.lich1.hide()

        # Khai báo Calendar 2
        self.lich2 = QtWidgets.QCalendarWidget(self)  # Khởi tạo màn hình Calendar2 khi ấn nút
        self.lich2.setWindowTitle("Calendar")
        self.lich2.move(1100, 400)
        self.lich2.resize(400,300)
        self.lich2.clicked[QtCore.QDate].connect(self.get_data2)
        self.lich2.hide()

    def get_data1(self, date):
        self.uic.dateEdit_1.setDate(date)       # Lấy giá trị ngày và giờ lịch 1
        self.lich1.hide()

    def get_data2(self, date):
        self.uic.dateEdit_2.setDate(date)       # Lấy giá trị ngày và giờ lịch 2
        self.lich2.hide()

    def start_capture_video(self):
        self.uic.pushButton_start_camera.setEnabled(False)
        self.uic.pushButton_stop_camera.setEnabled(True)
        self.thread[1] = capture_video(index=1)
        self.thread[1].start()
        self.thread[1].signal.connect(self.show_webcam)


    def stop_capture_video(self):
     # if 1 in self.thread:
        self.uic.pushButton_start_camera.setEnabled(True)
        self.uic.pushButton_stop_camera.setEnabled(False)
        self.thread[1].stop()

    def show_webcam(self, cv_img):
        """Cập nhật label với hình ảnh mới từ opencv"""
        cv_img = cv2.flip(cv_img, 1)  # Lật tấm hình lại để xử lý
        qt_img = self.convert_cv_qt(cv_img)
        self.uic.label_camera.setPixmap(qt_img)

    def convert_cv_qt(self, cv_img):
        rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(800, 600, QtCore.Qt.KeepAspectRatio)
        return QPixmap.fromImage(p)

    def start_home(self):  # Code để điều khiển Drone bay về vị trí ban đầu
        pass

    def show_map(self):
        pass

    def closeEvent(self, event):
        self.stop_capture_video()
        super().closeEvent(event)

    def choose_Ui_Form_1(self):  # Hàm mở giao diện Form (A_Few_Location)
        self.main_win2 = a_few_location()
        self.main_win2.show()

    def run_drone(self):
        pass

    def stop_drone(self):
        pass

    def reset(self):
        pass

    def start_mapping(self):
        pass

    def stop_mapping(self):
        pass

    def search(self):
        if self.uic.comboBox_chose_the_mode.currentIndex() == 1:          # Chọn mode xem Account
            start = self.uic.dateEdit_1.text()
            date_str = start
            # Định dạng chuỗi ban đầu
            date_format = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj = datetime.strptime(date_str, date_format)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str = datetime_obj.strftime("%Y-%m-%d")

            end = self.uic.dateEdit_2.text()
            date_str1 = end
            # Định dạng chuỗi ban đầu
            date_format1 = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj1 = datetime.strptime(date_str1, date_format1)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str1 = datetime_obj1.strftime("%Y-%m-%d")

            sql = "SELECT * FROM history_access WHERE DATE(Date) BETWEEN %s AND %s"
            mycursor.execute(sql, (new_date_str, new_date_str1))
            # Lấy tất cả các bản ghi từ kết quả truy vấn
            myresult = mycursor.fetchall()

            # Tạo table với số hàng và cột
            self.uic.tableWidget_account.setRowCount(len(myresult))
            self.uic.tableWidget_account.setColumnCount(4)

            # Thiết lập font chữ
            font = QFont()
            font.setBold(True)
            font.setPointSize(12)  # set font size to 12

            # Thiết lập tiêu đề, độ rộng các cột, hàng
            self.uic.tableWidget_account.setHorizontalHeaderLabels(['Date', 'Time', 'Username', 'Role'])
            self.uic.tableWidget_account.horizontalHeader().setFont(font)         # Set font cho tiêu đề ngang
            self.uic.tableWidget_account.verticalHeader().setVisible(False)       # Ẩn phần tiêu đề dọc

            for i in range(4):
                self.uic.tableWidget_account.setColumnWidth(i, 190)
            self.uic.tableWidget_account.setStyleSheet("QTableWidget {border: 3px solid black;}")
            self.uic.tableWidget_account.setFont(font)

            table_row = 0
            for row in myresult:
                date_str3 = row[0].strftime("%d/%m/%Y")
                self.uic.tableWidget_account.setItem(table_row, 0, QTableWidgetItem(str(date_str3)))
                self.uic.tableWidget_account.item(table_row, 0).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 1, QTableWidgetItem(str(row[1])))
                self.uic.tableWidget_account.item(table_row, 1).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 2, QTableWidgetItem(str(row[2])))
                self.uic.tableWidget_account.item(table_row, 2).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 3, QTableWidgetItem(row[3]))
                self.uic.tableWidget_account.item(table_row, 3).setTextAlignment(Qt.AlignCenter)
                table_row += 1

        elif self.uic.comboBox_chose_the_mode.currentIndex() == 2:       # Chọn mode xem Detect Box
            # # datetime object containing current date and time
            # now = datetime.now()
            # # Tách ngày và giờ
            # date = now.date()
            # time = now.time()
            # print(type(date))
            # print("date and time =", date)
            # query2 = "INSERT INTO history_access (Date, Time, Username, Role) VALUES (%s, %s, %s, %s)"
            # mycursor.execute(query2, (date, time, result[2], result[5]))
            # db.commit()

            start = self.uic.dateEdit_1.text()
            date_str = start
            # Định dạng của chuỗi ban đầu
            date_format = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj = datetime.strptime(date_str, date_format)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str = datetime_obj.strftime("%Y-%m-%d")

            end = self.uic.dateEdit_2.text()
            date_str1 = end
            # Định dạng của chuỗi ban đầu
            date_format1 = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj1 = datetime.strptime(date_str1, date_format1)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str1 = datetime_obj1.strftime("%Y-%m-%d")
            # start_date = '2023-04-01'
            # end_date = '2023-04-26'

            sql = "SELECT * FROM detect_box WHERE DATE(Date) BETWEEN %s AND %s"
            mycursor.execute(sql, (new_date_str, new_date_str1))
            # Lấy tất cả các bản ghi từ kết quả truy vấn
            myresult = mycursor.fetchall()

            # Tạo table với số hàng và cột
            self.uic.tableWidget_account.setRowCount(len(myresult))
            self.uic.tableWidget_account.setColumnCount(3)

            # Thiết lập font chữ
            font = QFont()
            font.setBold(True)
            font.setPointSize(12)  # set font size to 12

            # Thiết lập tiêu đề, độ rộng các cột, hàng
            self.uic.tableWidget_account.setHorizontalHeaderLabels(['Date', 'Time', 'QRCode'])
            self.uic.tableWidget_account.horizontalHeader().setFont(font)
            self.uic.tableWidget_account.verticalHeader().setVisible(False)

            for i in range(4):
                self.uic.tableWidget_account.setColumnWidth(i, 190)
            self.uic.tableWidget_account.setStyleSheet("QTableWidget {border: 3px solid black;}")
            self.uic.tableWidget_account.setFont(font)
            table_row = 0
            for row in myresult:
                date_str3 = row[1].strftime("%d/%m/%Y")
                self.uic.tableWidget_account.setItem(table_row, 0, QTableWidgetItem(str(date_str3)))
                self.uic.tableWidget_account.item(table_row, 0).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 1, QTableWidgetItem(str(row[2])))
                self.uic.tableWidget_account.item(table_row, 1).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 2, QTableWidgetItem(str(row[3])))
                self.uic.tableWidget_account.item(table_row, 2).setTextAlignment(Qt.AlignCenter)
                # self.uic.tableWidget_account.setItem(table_row, 3, QTableWidgetItem(row[4]))
                # self.uic.tableWidget_account.item(table_row, 3).setTextAlignment(Qt.AlignCenter)

                table_row += 1

            else:print("product")
    def clear(self):
        self.uic.tableWidget_account.clear()
        self.uic.tableWidget_account.setRowCount(0)
        self.uic.tableWidget_account.setColumnCount(0)

    def calendar1(self):
        if self.lich1.isHidden():
            self.lich1.show()
        else:
            self.lich1.hide()

    def calendar2(self):
        if self.lich2.isHidden():
            self.lich2.show()
        else:
            self.lich2.hide()

    def load_data(self):
        self.product_win = MainWindow_1()
        self.product_win.show()

    def report(self):
        if self.uic.comboBox_chose_the_mode.currentIndex()==0:
            try:
                mycursor.execute("SELECT * FROM qrcodes_info")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("Information of Products.xlsx")
                db.commit()

                QMessageBox.information(self,"Message"," Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()

        if self.uic.comboBox_chose_the_mode.currentIndex() == 1:
            try:
                mycursor.execute("SELECT * FROM history_access")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("History Access.xlsx")
                db.commit()

                QMessageBox.information(self, "Message", " Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()

        if self.uic.comboBox_chose_the_mode.currentIndex() == 2:
            try:
                mycursor.execute("SELECT * FROM detect_box")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("Detected Box.xlsx")
                db.commit()

                QMessageBox.information(self, "Message", " Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()
class MainWindow_1(QWidget):               # Form products_info
    def __init__(self):
        super().__init__()
        self.uic = Ui_Form_2()  # Tạo đối tượng UI
        self.uic.setupUi(self)  # Thiết lập giao diện

        # Create a database connection object
        self.db = ConnectDatabase()

        # Connect UI elements to class variables
        self.qrcode_id = self.uic.lineEdit
        # # Restrict input to integers
        # self.qrcode_id.setValidator(QIntValidator)

        self.location = self.uic.lineEdit_2
        self.quantity = self.uic.lineEdit_3
        self.type = self.uic.lineEdit_4
        self.status = self.uic.lineEdit_5
        self.date = self.uic.lineEdit_6

        self.add_btn = self.uic.add_btn
        self.update_btn = self.uic.update_btn
        self.select_btn = self.uic.select_btn
        self.search_btn = self.uic.search_btn
        self.clear_btn = self.uic.clear_btn
        self.delete_btn = self.uic.delete_btn

        self.result_table = self.uic.tableWidget
        self.result_table.setSortingEnabled(False)
        self.buttons_list = self.uic.function_frame.findChildren(QPushButton)

        # Initialize signal-slot conections
        self.init_signal_slot()

        # Populate the initial data in the table and Type of goods/ Status dropdowns
        self.search_info()
        # self.update_type_status()

    def init_signal_slot(self):
        # Connect buttons to their respective functions
        self.add_btn.clicked.connect(self.add_info)
        self.search_btn.clicked.connect(self.search_info)
        self.clear_btn.clicked.connect(self.clear_form_info)
        self.select_btn.clicked.connect(self.select_info)
        self.update_btn.clicked.connect(self.update_info)
        self.delete_btn.clicked.connect(self.delete_info)

    # def update_type_status(self):
    #     # Function to populate the type and status dropdowns
    #     type_result = self.db.get_all_types()
    #     status_result = self.db.get_all_status()
    #
    #     self.type.clear()
    #     self.status.clear()
    #
    #     type_list =[""]
    #     for item in type_result:
    #         for k,v in item.items():
    #             if v != "":
    #                 type_list.append(v)
    #
    #     status_list = [""]
    #     for item in status_result:
    #         for k,v in item.items():
    #             if v != "":
    #                 status_list.append(v)
    #
    #     if len(type_list) > 1:
    #         self.type.addItem(type_list)
    #
    #     if len(status_list) > 1:
    #         self.status.addItem(status_list)

    def search_info(self):
        # Function to search for qrcode information and populate the table
        # self.update_type_status()
        qrcode_info = self.get_qrcode_info()

        qrcode_result = self.db.search_info(
            qrcode_id=str(qrcode_info["qrcode_id"]),
            location=qrcode_info["location"],
            quantity=qrcode_info["quantity"],
            date=qrcode_info["date"],
            type=qrcode_info["type"],
            status=qrcode_info["status"],
        )

        self.show_data(qrcode_result)





    def add_info(self):
        # Function to add qrcode information
        self.disable_buttons()

        qrcode_info = self.get_qrcode_info()

        if qrcode_info["qrcode_id"] and qrcode_info["location"]:
            check_result = self.check_qrcode_id(qrcode_id=str(qrcode_info["qrcode_id"]))    # Kiểm tra xem đã tồn tại qrcode hay chưa

            if check_result:
                QMessageBox.information(self,"Warning","Please input a new QR Code ID",QMessageBox.StandardButton.Ok)
                self.enable_buttons()
                return

            add_result = self.db.add_info(qrcode_id=str(qrcode_info["qrcode_id"]),
                                          location=qrcode_info["location"],
                                          quantity=qrcode_info["quantity"],
                                          date=qrcode_info["date"],
                                          type=qrcode_info["type"],
                                          status=qrcode_info["status"],
                                          )
            if add_result:
                QMessageBox.information(self,"Warning",f"Add fail:{add_result}, Please try again.",QMessageBox.StandardButton.Ok)

        else:
            QMessageBox.Information(self,"Warning","Please input QR Code ID and location",QMessageBox.StandardButton.Ok)

        self.search_info()
        self.enable_buttons()
    def clear_form_info(self):
        # Function to clear the form
        # self.update_type_status()
        self.qrcode_id.clear()
        self.qrcode_id.setEnabled(True)
        self.location.clear()
        self.quantity.clear()
        self.type.clear()
        self.status.clear()
        self.date.clear()


    def update_info(self):
        # Function to update qrcode information
        new_qrcode_info = self.get_qrcode_info()

        if new_qrcode_info["qrcode_id"]:
            update_result = self.db.update_info(
                qrcode_id=str(new_qrcode_info["qrcode_id"]),
                location=new_qrcode_info["location"],
                quantity=new_qrcode_info["quantity"],
                date=new_qrcode_info["date"],
                type=new_qrcode_info["type"],
                status=new_qrcode_info["status"],
            )

            if update_result:
                QMessageBox.information(self,"Warning", f"Fail to update the information:{update_result}, Please try again.",
                                        QMessageBox.StandardButton.Ok)
            else:
                self.search_info()

        else:
            QMessageBox.information(self,"Warning", "Please select one qrcode information to update.")


    # def delete_info(self):
    #     # Function to delete qrcode information
    #     select_row = self.result_table.currentRow()     # Nếu không có hàng nào được chọn thì index = -1
    #     if select_row != -1:      # Kiểm tra xem có hàng nào được chọn hay không (!= -1 : tức là có 1 hàng được chọn )
    #         select_option = QMessageBox.warning(self,"Warning", "Are you sure to delete it?",
    #                                             QMessageBox.StandardButton.Yes| QMessageBox.StandardButton.No)
    #         if select_option == QMessageBox.StandardButton.Yes:
    #             qrcode_id = self.result_table.item(select_row,0).text().strip()
    #
    #             delete_result = self.db.delete_info(qrcode=qrcode_id)
    #             if not delete_result:
    #                 self.search_info()
    #             else:
    #                 QMessageBox.information(self,"Warning", f"Fail to delete the information: {delete_result},please try again",
    #                                         QMessageBox.StandardButton.Ok)
    #         # else:
    #         #     select_option == QMessageBox.StandardButton.Cancel
    #         #     self.search_info()
    #     else:
    #         QMessageBox.information(self,"Warning","Please select one qrcode information to delete",QMessageBox.StandardButton.Ok)

    def delete_info(self):
        select_row = self.result_table.currentRow()
        if select_row != -1:
            select_option = QMessageBox.warning(self, "Warning", "Are you sure to delete it?",
                                                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
            if select_option == QMessageBox.StandardButton.Ok:
                qrcode_id = self.result_table.item(select_row, 0).text().strip()
                print(f"Deleting QR Code ID: {qrcode_id}")
                delete_result = self.db.delete_info(qrcode_id=qrcode_id)
                print(f"Delete Result: {delete_result}")
                if not delete_result:
                    self.search_info()
                else:
                    QMessageBox.information(self, "Warning",
                                            f"Fail to delete the information: {delete_result}, please try again",
                                            QMessageBox.StandardButton.Ok)
            else:
                self.search_info()
        else:
            QMessageBox.information(self, "Warning", "Please select one qrcode information to delete",
                                    QMessageBox.StandardButton.Ok)

    def select_info(self):
        # Function to select and populate student information in the form
        select_row = self.result_table.currentRow()
        if select_row != -1:      # Nếu có hàng được chọn
            self.qrcode_id.setEnabled(False)
            qrcode_id = self.result_table.item(select_row,0).text().strip()
            location = self.result_table.item(select_row, 1).text().strip()
            quantity = self.result_table.item(select_row, 2).text().strip()
            type = self.result_table.item(select_row, 3).text().strip()
            status = self.result_table.item(select_row, 4).text().strip()
            date = self.result_table.item(select_row, 5).text().strip()

            self.qrcode_id.setText(qrcode_id)
            self.location.setText(location)
            self.quantity.setText(quantity)
            self.type.setText(type)
            self.status.setText(status)
            self.date.setText(date)

        else:
            QMessageBox.information(self,"Warning", "Please select one qrcode information",
                                    QMessageBox.StandardButton.Ok)


    def disable_buttons(self):
        # Disable all the buttons
        for button in self.buttons_list:
            button.setDisabled(True)

    def enable_buttons(self):
        # enable all the buttons
        for button in self.buttons_list:
            button.setDisabled(False)

    def get_qrcode_info(self):
        # Function to retrieve qrcode information from the form
        qrcode_id = self.qrcode_id.text().strip()
        location = self.location.text().strip()
        quantity = self.quantity.text().strip()
        date = self.date.text().strip()
        type = self.type.text().strip()
        status = self.status.text().strip()

        qrcode_info = {
            "qrcode_id": qrcode_id,
            "location": location,
            "quantity": quantity,
            "date": date,
            "type": type,
            "status": status,
        }

        return qrcode_info

    def check_qrcode_id(self,qrcode_id):
        # Function to check if a qrcode_id already exists
        result = self.db.search_info(qrcode_id=qrcode_id)

        return result

    def show_data(self,result):    # Hiển thị dữ liệu trong bảng kết quả
        # Function to populate the table with qrcode information
        if result:
            self.result_table.setRowCount(0)
            self.result_table.setRowCount(len(result))

            for row, info in enumerate(result):
                info_list =[
                    info["qrcode"],
                    info["location"],
                    info["quantity"],
                    info["type"],
                    info["status"],
                    info["date"],
                ]

                for column, item in enumerate(info_list):
                    cell_item = QTableWidgetItem(str(item))
                    self.result_table.setItem(row,column,cell_item)
        else:
            self.result_table.setRowCount(0)
            return

class a_few_location(QWidget):
    def __init__(self):
        super().__init__()
        self.uic2 = Ui_Form_1()  # Tạo giao diện a_few_location
        self.uic2.setupUi(self)  # Thiết lập giao diện

class capture_video(QThread):
    signal = pyqtSignal(np.ndarray)
    def __init__(self, index=0):
        super().__init__()
        self.index = index
        print("Start threading capture video", self.index)
        super(capture_video,self).__init__()

    def run(self):
        cap = cv2.VideoCapture(0)
        while True:
            ret, cv_img = cap.read()
            if ret:
                self.signal.emit(cv_img)  # cho phép truyền tín hiệu ra

    def stop(self):
        print("stop threading capture video", self.index)
        self.terminate()  # Dừng thread

class capture_map(QThread):
    signal = pyqtSignal(np.ndarray)

    def __init__(self, index=0):
        super().__init__()
        self.index = index
        print("Start threading capture map", self.index)

    def run(self):
        # Điền code
        self.signal.emit()

    def stop(self):
        print("stop threading capture map", self.index)
        self.terminate()  # Dừng thread

if __name__ == "__main__":
    # Chạy ứng dụng
    app = QApplication(sys.argv)
    login_win = LoginWindow()
    login_win.show()
    sys.exit(app.exec())